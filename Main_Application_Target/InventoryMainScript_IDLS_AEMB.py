'''
Project Name: InventoryEXT_SMV_GUI - Inventory To Simplified Model View, Code Material Design
Proper Name: Excel Inventory To Simplified View
Version: unknown.beta
Authors: Janrey Tuazon Licas - App Design (Semi)
        Sam Matienzo (Database Worker)
GUI Library Used: Kivy
    Base Theme: KivyMD (Material Design Preset Designs)

Date Initialization: 02/21/2019

'''
from kivy.app import App
from kivy.lang import Builder
from kivy.metrics import dp
from kivy.core.window import Window
from kivy.properties import ObjectProperty
from kivy.uix.image import Image
from kivy.uix.widget import Widget
from kivy.properties import ObjectProperty, NumericProperty, StringProperty
from kivy.uix.screenmanager import ScreenManager, Screen, FadeTransition
from kivy.config import Config
from kivy.logger import Logger as LoggerDebug
from kivymd.bottomsheet import MDListBottomSheet, MDGridBottomSheet
from kivymd.button import MDIconButton
from kivymd.date_picker import MDDatePicker
from kivymd.dialog import MDDialog
from kivymd.label import MDLabel
from kivymd.list import ILeftBody, ILeftBodyTouch, IRightBodyTouch, BaseListItem
from kivymd.material_resources import DEVICE_TYPE
from kivymd.navigationdrawer import MDNavigationDrawer, NavigationDrawerHeaderBase, NavigationDrawerIconButton
from kivymd.selectioncontrols import MDCheckbox
from kivymd.snackbar import Snackbar
from kivymd.theming import ThemeManager
from kivymd.time_picker import MDTimePicker
from kivymd.toolbar import Toolbar
from kivymd.tabs import MDTab
from kivymd.textfields import MDTextField
from kivy.utils import get_color_from_hex
from kivymd.color_definitions import colors
from kivymd.list import OneLineListItem
from kivymd.list import TwoLineListItem
from kivymd.list import ThreeLineListItem
from kivymd.list import OneLineAvatarListItem
from kivymd.list import OneLineIconListItem
from kivymd.list import OneLineAvatarIconListItem
from kivy.clock import Clock
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
import sqlite3
#Clock.max_iteration = 20
#class HackedDemoNavDrawer(MDNavigationDrawer):
#    # DO NOT USE
#    def add_widget(self, widget, index=0):
#        if issubclass(widget.__class__, BaseListItem):
#            self._list.add_widget(widget, index)
#            if len(self._list.children) == 1:
#                widget._active = True
#                self.active_item = widget
#            # widget.bind(on_release=lambda x: self.panel.toggle_state())
#            widget.bind(on_release=lambda x: x._set_active(True, list=self))
#        elif issubclass(widget.__class__, NavigationDrawerHeaderBase):
#            self._header_container.add_widget(widget)
#        else:
#            super(MDNavigationDrawer, self).add_widget(widget, index)


class MD_InventoryEXI_SMV_GUI(App):
    theme_cls = ThemeManager()
    previous_date = ObjectProperty()
    #title = "Welcome to Excel Inventory To Simplified Model View | EXI_SMV" #Excel to Simplified Model VIew in GUI
    title = "Excel Inventory To Simplified Model View | EXI_SMV - Individual Use" #Excel to Simplified Model VIew in GUI
    Icon_Status = 'brightness-1'
    def KivyConfig_Init(self): # Get Values from DB and Change App Behavior Accordingly
        pass

    def KivyConfig_Setup_FirstTime(self): #This Config can be changed from DB
        #Config.set('graphics', 'fullscreen', 'fake')
        Config.set('graphics','width','1024')
        Config.set('graphics','height','768')
        Config.set('graphics','show_cursor','1')
        Config.set('graphics','minimum_width','1024')
        Config.set('graphics','minimum_height','768')
        Config.set('kivy','exit_on_escape','0')
        Config.write()
    def SQLite_Generate_FirstTime(self):
        pass

    def SQLite_Init(self): #Init Files and Creates when it doesnt exist but show snackbar that is doesnt exist
        pass

    def SQLite_ChangeModify(self): # Not Sure
        pass

    def CriticalComponent_Check(self):
        return 1
        if self.KivyConfig_Init() == 'Passed':
            LoggerDebug.info()
        elif self.KivyConfig_Init() == 'FileDoesNotExist': #Not Consistent, Maybe Clarify this one soon
            LoggerDebug.info('FILE EXISTENT: App knows this is not the first. File does not exist. Relocate the application...') # Add save file
            raise Exception('FILE EXISTENT: App knows this is not the first. File does not exist. Relocate the application...')
            self.get_running_app().stop() 
        else: #Add Path does not exist
            if self.KivyConfig_Setup_FirstTime() == 'Passed':
                LoggerDebug.info('Kivy Config: First Time Settings is loaded and saved.') # Add save file
                if self.SQLite_Init() == 'Passed':
                    LoggerDebug.info('Kivy Config: First Time Settings is loaded and saved.') # Add save file
                else:
                    if self.SQLite_Generate_FirstTime() == 'Passed':
                        LoggerDebug.info('Kivy Config: First Time Files is Initialized and Ready To Go!') # Add save file
                        return 1
                    else:
                        return 0
            else:
                return 0

    def build(self):
        self.MainClassBuildFile = Builder.load_file("MD_DesignClass_File.kv")
        self.theme_cls.primary_palette = 'DeepOrange'
        self.theme_cls.primary_hue = '400'
        self.theme_cls.theme_style = 'Light'
        on_dropfile=self._on_file_drop
        return self.MainClassBuildFile
        
    def _on_file_drop(self, window, file_path):
        print(file_path)
        return
    
    def MD_DataManage_ShowCellToList(self):
        pass

    def MD_DataManage_Modify(self):
        pass

    def MD_FirstTime_DataSubmission(self):
        pass

    def MDCard_GenerateReport(self):
        pass

    #For Add and Delete
    def MDAuthor_DeleteData(self):
        pass

    def MDAuthor_AddData(self):
        pass

    def MDAuthor_EditData(self):
        pass

    def MDStartup_FirstTime_AdminAccess(self, BooleanPrompt):
        pass
    
    #For Checking Previledge on Accessing Admin Features
    def MDAuthor_CheckAdminPrev_OnAction(self):
        pass

    # For Logon Screen
    def MDAuthor_DialogBooleanPrompt(self):
        pass

    #For Logon and Admin for Modifying User and Assign
    def MDAuthor_PasswordAccessPrompt(self):
        pass
        
    def MDToolbar_DynamicContent_Change(self, params_title, params_pallete, params_left_action_items, params_right_action_items):
        self.root.ids.ToolbarMain.title = params_title
        self.root.ids.ToolbarMain.md_bg_color = get_color_from_hex(colors[params_pallete][self.theme_cls.primary_hue])
        self.root.ids.ToolbarMain.background_palette = params_pallete
        self.root.ids.ToolbarMain.left_action_items = params_left_action_items
        self.root.ids.ToolbarMain.right_action_items = params_right_action_items

    #For First TIme

    def MDNavigationDrawer_DynamicContent_Change(self, Params_Mode):
        if Params_Mode == 'User.Logon_Passed':
            #self.root.ids.
            pass
        elif Params_Mode == 'User.LogonPrompt':
            pass
        elif Params_Mode == 'User.PromptedAdmin_Passed':
            pass
        else:
            LoggerDebug.error('RUNTIME ERROR: Unknown String Passed with No Valid Statement To Run!')
            raise ValueError('RUNTIME ERROR:  Unknown String Passed with No Valid Statement To Run!')
    
    def MDButton_DarkMode(self):
        if self.root.ids.ToolbarMain.darkBooleanParameter == True and self.theme_cls.theme_style == 'Dark':
            self.root.ids.ToolbarMain.darkBooleanParameter = False
            self.root.ids.ToolbarMain.right_action_items =  [['brightness-2', lambda x: MD_InventoryEXI_SMV_GUI.MDButton_DarkMode(self)]]
            self.theme_cls.theme_style = 'Light'
            # Make it time based, if possible
            self.MDUserNotif_SnackbarHandler('Dark Mode Activated!')
        elif self.root.ids.ToolbarMain.darkBooleanParameter == False and self.theme_cls.theme_style == 'Light':
            self.root.ids.ToolbarMain.darkBooleanParameter = True
            self.root.ids.ToolbarMain.right_action_items =  [['brightness-7', lambda x: MD_InventoryEXI_SMV_GUI.MDButton_DarkMode(self)]]
            self.theme_cls.theme_style = 'Dark'
            self.MDUserNotif_SnackbarHandler('Light Mode Activated!')
        else:
            LoggerDebug.error('MDButton_DarkMode received invalid parameters, reset SQL Database or modify the property of App_ReadibilityMode to Light or Dark in String Form')
            raise ValueError('MDButton_DarkMode received invalid parameters, reset SQL Database or modify the property of App_ReadibilityMode to Light or Dark in String Form')
    
    def MDUserNotif_SnackbarHandler(self, params_message):
        Snackbar(text=params_message).show()

    def ExcelLoadInit_Data_Editor(self, ExcelLoad_Method):
        #Should be selectable
        #ExcelID_UniqueListIndex = 0
        ListEntry_Inf = []
        ListCount = 0
        CounterCheck = 1
        CounterList = 1
        StringAppend = 0
        Content_ListElement = 0
        Content_ListElement_Deletion = 0
        Content_DeletionList = 1
        Content_CompletionList = 1
        ExcelContainer_Val_ItemNumber = ''
        ExcelContainer_Text_ParticularName = ''
        ExcelContainer_Val_OnHand = ''
        ExcelContainer_Val_Proposed = ''
        ExcelContainer_CostUnit_Type = ''
        ExcelContainer_CostUnit_Val = ''
        ExcelContainer_CostUnit_Total = ''
        ExcelContainer_Quarter_Allot_1 = ''
        ExcelContainer_Quarter_Allot_2 = ''
        ExcelContainer_Quarter_Allot_3 = ''
        ExcelContainer_Quarter_Allot_4 = ''
        Concatenation_PrimaryText = ''
        Concatenation_SecondaryText = ''
        Start_CellPoint = ''
        End_CellPoint = ''
        #ExcelID_MDTabbedPanel = 0 This was intended to use on multi-user version along with commented function below
        try:
            ExcelFile = load_workbook(self.root.ids.DataBind_FilePath.text, data_only = self.root.ids.DataBind_FileArguments.text)
            ExcelWorksheet = ExcelFile['{0}'.format(self.root.ids.DataBind_ExcelFileSheet.text)]
            
            if ExcelLoad_Method == 'Unload' or ExcelLoad_Method == 'Reload':
                for Element_InList in range(1, 101):
                    print('[Excel Data Unload / Reload] ListElement # ', Element_InList)
                    self.root.ids['ListElement_%d' % Element_InList].cellStart = '0'
                    self.root.ids['ListElement_%d' % Element_InList].cellEnd = '0'
                    self.root.ids['ListElement_%d' % Element_InList].text = '#??? - <No Data>'
                    self.root.ids['ListElement_%d' % Element_InList].secondary_text = 'There is no data to load...'
                    self.root.ids['ListElement_%d' % Element_InList].disabled = True
                    self.root.ids['ListElement_%d' % Element_InList].itemNumber = '0'
                    self.root.ids['ListElement_%d' % Element_InList].particularName = '0'
                    self.root.ids['ListElement_%d' % Element_InList].onHandVal = '0'
                    self.root.ids['ListElement_%d' % Element_InList].proposedVal = '0'
                    self.root.ids['ListElement_%d' % Element_InList].unitType = '0'
                    self.root.ids['ListElement_%d' % Element_InList].unitVal = '0'
                    self.root.ids['ListElement_%d' % Element_InList].quarterOne = '0'
                    self.root.ids['ListElement_%d' % Element_InList].quarterTwo = '0'
                    self.root.ids['ListElement_%d' % Element_InList].quarterThree = '0'
                    self.root.ids['ListElement_%d' % Element_InList].quarterFour = '0'
                    self.root.ids['ListElement_%d' % Element_InList].totalVal = '0'
                    self.root.ids.Resource_ItemNumVal.text = ''
                    self.root.ids.Resource_ParticularProperty.text = ''
                    self.root.ids.Resource_OnHandVal.text = ''
                    self.root.ids.Resource_ProposedVal.text = ''
                    self.root.ids.Resource_UnitTypeProperty.text = ''
                    self.root.ids.Resource_UnitVal.text = ''
                    self.root.ids.Quartile_TextFieldVal_One.text = ''
                    self.root.ids.Quartile_TextFieldVal_Two.text = ''
                    self.root.ids.Quartile_TextFieldVal_Three.text = ''
                    self.root.ids.Quartile_TextFieldVal_Four.text = ''
                    self.root.ids.Total_ComputedVal.text = ''
                self.root.ids.FileOpt_Progress_SaveImport.disabled = True
                self.root.ids.FileOpt_Progress_SaveExport.disabled = True
                self.root.ids.FileOpt_ReInit_ReloadDataFile.disabled = True
                self.root.ids.FileOpt_EndInst_UnloadDataFile.disabled = True
                self.root.ids.FileOpt_FirstInst_LoadDataFile.disabled = False
                self.MDUserNotif_SnackbarHandler("Data is Unbinded from the Editor.")
                print("[DATA - EXCEL, ACTION] Data is Unbinded from the Editor.")
            if ExcelLoad_Method == 'Load' or ExcelLoad_Method == 'Reload':
                for EachDataSheet in ExcelWorksheet.iter_rows(min_row=int(self.root.ids.DataBind_RowCellStartPoint.text), max_row=int(self.root.ids.DataBind_RowCellEndPoint.text), min_col=column_index_from_string(self.root.ids.DataBind_ColumnCellStartPoint.text), max_col=column_index_from_string(self.root.ids.DataBind_ColumnCellEndPoint.text)):
                    # This could be changed, or prolly make the set of ids into list, not dictionary
                    # This would loop for each element...
                    for cell in EachDataSheet:
                        '''
                        I can create those widgets dynamically but not with IDS
                        I would rather go to Iterate Through IDS of the Statically Created ListItem Method than
                        trying to use weakref and other such methods when in fact I will just get myself go to the next problem such as 
                        accessing ids in runtime without updating self.ids... It is not updatable... Since kivy store those ids in self.ids without update
                        even adding a wdiget. So since, the project deadlien is almost ahead of time. We would rather go to this static method, it's really hard to accept
                        as a Lead Developer of this application but since time goes by, we have to implement something even though this is one of the most important component in the 
                        entire application. In the recent template, there are spcaes that needs to be avoided...
                        '''
                        #print(cell, cell.value, CounterCheck)
                        Content_ListElement += 1
                        print('[Excel Data Load] Content Loop Through Inside ', Content_ListElement, 'Cell Stats', cell,'| Cell Value', cell.value)
                        if Content_ListElement == 1 and Content_ListElement <= 15:
                            if cell.value is None:
                                ExcelContainer_Val_ItemNumber = '-'
                                Start_CellPoint = '{0}{1}'.format(cell.column, cell.row)
                                self.root.ids['ListElement_%d' % Content_CompletionList].cellStart = Start_CellPoint

                            else:
                                ExcelContainer_Val_ItemNumber = cell.value
                                Start_CellPoint = '{0}{1}'.format(cell.column, cell.row)
                                self.root.ids['ListElement_%d' % Content_CompletionList].cellStart = Start_CellPoint

                        elif Content_ListElement == 2 and Content_ListElement <= 15:
                                continue
                        elif Content_ListElement == 3 and Content_ListElement <= 15:
                            if cell.value is None:
                                ExcelContainer_Text_ParticularName = '<No Description>'
                                print(ExcelContainer_Text_ParticularName)

                            else:
                                ExcelContainer_Text_ParticularName = cell.value
                                print('Output',ExcelContainer_Text_ParticularName)

                        elif Content_ListElement == 4 and Content_ListElement <= 15:
                                continue
                        elif Content_ListElement == 5 and Content_ListElement <= 15:
                            if cell.value is None:
                                ExcelContainer_Val_OnHand = 'None'

                            else:
                                ExcelContainer_Val_OnHand = cell.value

                        elif Content_ListElement == 6 and Content_ListElement <= 15:
                                continue
                        elif Content_ListElement == 7 and Content_ListElement <= 15:
                            if cell.value is None:
                                ExcelContainer_Val_Proposed = 'None'

                            else:
                                ExcelContainer_Val_Proposed = cell.value

                        elif Content_ListElement == 8 and Content_ListElement <= 15:
                                continue
                        elif Content_ListElement == 9 and Content_ListElement <= 15:
                            if cell.value is None:
                                ExcelContainer_CostUnit_Type = 'Unknown'

                            else:
                                ExcelContainer_CostUnit_Type = cell.value

                        elif Content_ListElement == 10 and Content_ListElement <= 15:
                            if cell.value is None:
                                ExcelContainer_CostUnit_Val = '0.00'

                            else:
                                ExcelContainer_CostUnit_Val = cell.value

                        elif Content_ListElement == 11 and Content_ListElement <= 15:
                            if cell.value is None:
                                ExcelContainer_CostUnit_Total = '0.00'

                            else:
                                ExcelContainer_CostUnit_Total = cell.value

                        elif Content_ListElement == 12 and Content_ListElement <= 15:
                            if cell.value is None:
                                ExcelContainer_Quarter_Allot_1 = '-'

                            else:
                                ExcelContainer_Quarter_Allot_1 = cell.value

                        elif Content_ListElement == 13 and Content_ListElement <= 15:
                            if cell.value is None:
                                ExcelContainer_Quarter_Allot_1 = '-'

                            else:
                                ExcelContainer_Quarter_Allot_2 = cell.value

                        elif Content_ListElement == 14 and Content_ListElement <= 15:
                            if cell.value is None:
                                ExcelContainer_Quarter_Allot_4 = '-'

                            else:
                                ExcelContainer_Quarter_Allot_3 = cell.value

                        elif Content_ListElement == 15 and Content_ListElement <= 15:
                            if cell.value is None:
                                ExcelContainer_Quarter_Allot_4 = '-'
                                Content_ListElement = 0
                                End_CellPoint = '{0}{1}'.format(cell.column, cell.row)
                                self.root.ids['ListElement_%d' % Content_CompletionList].cellEnd = End_CellPoint
                                break
                            else:
                                ExcelContainer_Quarter_Allot_4 = cell.value
                                Content_ListElement = 0
                                End_CellPoint = '{0}{1}'.format(cell.column, cell.row)
                                self.root.ids['ListElement_%d' % Content_CompletionList].cellEnd = End_CellPoint
                                break
                    #print('Final Output',self.Concatenation_PrimaryText)
                    #Concatenation_PrimaryText = '{0} | {1}'.format(ExcelContainer_Val_ItemNumber,ExcelContainer_Text_ParticularName)
                    #Concatenation_SecondaryText = 'OnHand: {0}, Proposed:{1}, Unit Type:{2}, Unit Value:{3}, TotalVal:{4}, Q1:{6}, Q2:{7}, Q3:{8}, Q4:{9}'.format(ExcelContainer_Val_OnHand, ExcelContainer_Val_Proposed, ExcelContainer_CostUnit_Type, ExcelContainer_CostUnit_Val, ExcelContainer_CostUnit_Total, ExcelContainer_Quarter_Allot_1, ExcelContainer_Quarter_Allot_2, ExcelContainer_Quarter_Allot_3, ExcelContainer_Quarter_Allot_4)
                    self.root.ids['ListElement_%d' % Content_CompletionList].text = '{0} | {1}'.format(ExcelContainer_Val_ItemNumber,ExcelContainer_Text_ParticularName)
                    self.root.ids['ListElement_%d' % Content_CompletionList].secondary_text = 'OnHand: {0}, Proposed:{1}, Unit Type:{2}, Unit Value:{3}, TotalVal:{4}, Q1:{5}, Q2:{6}, Q3:{7}, Q4:{8}'.format(ExcelContainer_Val_OnHand, ExcelContainer_Val_Proposed, ExcelContainer_CostUnit_Type, ExcelContainer_CostUnit_Val, ExcelContainer_CostUnit_Total, ExcelContainer_Quarter_Allot_1, ExcelContainer_Quarter_Allot_2, ExcelContainer_Quarter_Allot_3, ExcelContainer_Quarter_Allot_4)
                    self.root.ids['ListElement_%d' % Content_CompletionList].disabled = False
                    #Save Data Here, because we cannot get that value on press when clicked...
                    self.root.ids['ListElement_%d' % Content_CompletionList].itemNumber = str(ExcelContainer_Val_ItemNumber)
                    self.root.ids['ListElement_%d' % Content_CompletionList].particularName = str(ExcelContainer_Text_ParticularName)
                    self.root.ids['ListElement_%d' % Content_CompletionList].onHandVal = str(ExcelContainer_Val_OnHand)
                    self.root.ids['ListElement_%d' % Content_CompletionList].proposedVal = str(ExcelContainer_Val_Proposed)
                    self.root.ids['ListElement_%d' % Content_CompletionList].unitType = str(ExcelContainer_CostUnit_Type)
                    self.root.ids['ListElement_%d' % Content_CompletionList].unitVal = str(ExcelContainer_CostUnit_Val)
                    self.root.ids['ListElement_%d' % Content_CompletionList].quarterOne = str(ExcelContainer_Quarter_Allot_1)
                    self.root.ids['ListElement_%d' % Content_CompletionList].quarterTwo = str(ExcelContainer_Quarter_Allot_2)
                    self.root.ids['ListElement_%d' % Content_CompletionList].quarterThree = str(ExcelContainer_Quarter_Allot_3)
                    self.root.ids['ListElement_%d' % Content_CompletionList].quarterFour = str(ExcelContainer_Quarter_Allot_4)
                    self.root.ids['ListElement_%d' % Content_CompletionList].totalVal = str(ExcelContainer_CostUnit_Total)
                    Content_CompletionList += 1
                self.MDUserNotif_SnackbarHandler("File Data Load Success!")
                if ExcelLoad_Method == 'Load':
                    self.root.ids.FileOpt_Progress_SaveImport.disabled = False
                    self.root.ids.FileOpt_Progress_SaveExport.disabled = False
                    self.root.ids.FileOpt_FirstInst_LoadDataFile.disabled = True
                    self.root.ids.FileOpt_ReInit_ReloadDataFile.disabled = False
                    self.root.ids.FileOpt_EndInst_UnloadDataFile.disabled = False
                #elif ExcelLoad_Method == 'Reload':
                    
        except:
            self.MDUserNotif_SnackbarHandler("File Data Load Failed Succesfully... Check the files or the arguments you set on the Data Import Data / Export Pull Out Section!")
                    # This value which instantly changed when its applied or when editing is done... (I guess)
                    # The value of this show not be saved in the way it is modified here. Get the formmula by getting the function reference from undo function
                    # Create a function to dice the formula and compute it with possible iterations from any n number
        #self.root.ids.MDList_UserInsertion_Selection.add_widget(TwoLineListItem(id=('IterationItem_%d' % CounterList), name=cell.value))
    def ExcelLoadDataTextField_OnClick(self, ListElement_Reference):
        self.root.ids.Resource_ItemNumVal.text = self.root.ids['%s' % ListElement_Reference].itemNumber
        self.root.ids.Resource_ParticularProperty.text = self.root.ids['%s' % ListElement_Reference].particularName
        self.root.ids.Resource_OnHandVal.text = self.root.ids['%s' % ListElement_Reference].onHandVal
        self.root.ids.Resource_ProposedVal.text = self.root.ids['%s' % ListElement_Reference].proposedVal
        self.root.ids.Resource_UnitTypeProperty.text = self.root.ids['%s' % ListElement_Reference].unitType
        self.root.ids.Resource_UnitVal.text = self.root.ids['%s' % ListElement_Reference].unitVal
        self.root.ids.Quartile_TextFieldVal_One.text = self.root.ids['%s' % ListElement_Reference].quarterOne
        self.root.ids.Quartile_TextFieldVal_Two.text = self.root.ids['%s' % ListElement_Reference].quarterTwo
        self.root.ids.Quartile_TextFieldVal_Three.text = self.root.ids['%s' % ListElement_Reference].quarterThree
        self.root.ids.Quartile_TextFieldVal_Four.text = self.root.ids['%s' % ListElement_Reference].quarterFour
        self.root.ids.Total_ComputedVal.text = self.root.ids['%s' % ListElement_Reference].totalVal

    def ExcelFile_SaveOnCurrentPath(self):
        ExcelFile = load_workbook(self.root.ids.DataBind_FilePath.text, data_only = self.root.ids.DataBind_FileArguments.text)
        ws = ExcelFile.active
        ws['A2'] = 'Tom'
        ws['B2'] = 30
        ws['B6'] = 30123

        ws['A3'] = 'Marry'
        ws['B3'] = 29
        # Save the file
        ExcelFile.save("sample.xlsx")
            #try:
        #    ExcelFile_Init = Workbook()
        #    
        #    ExcelFile_Init.save(self.root.ids.DataBind_FilePath.text)
        #except:
        #    self.MDUserNotif_SnackbarHandler('')
    def ExcelFile_SaveOnExportPath(self):
        try:
            ExcelFile_Init = Workbook()
            ExcelFile_Init.save(self.root.ids.DataBind_FileExportPath.text)
        except:
            self.MDUserNotif_SnackbarHandler('')

    def MDWorksheetWorker_DataEditor_Apply(self):
        # Create an Undo Method here?
        self.MDUserNotif_SnackbarHandler('Selected Data Edit has been applied!')

    def MDWorksheetWorker_DataEditor_ValueRaiseLowCaller(self, TextField_ID, ListElement_ItemNumber, TypeAction_ValueChange):
        try:
            Placement_Val_Conversion = self.root.ids.TextField_ID.text
            #self.root.ids.
            if TypeAction_ValueChange == 'Increment':
                Placement_Val_Conversion = int(Placement_Val_Conversion) + 1
                self.root.ids['%s' % TextField_ID].text = Placement_Val_Conversion
                #self.root.ids.['ListElement_%d' % ListElement_ItemNumber].
            elif TypeAction_ValueChange == 'Decrement':
                Placement_Val_Conversion = int(Placement_Val_Conversion) - 1
                self.root.ids.TextField_ID.text = Placement_Val_Conversion
                #self.root.ids.['ListElement_%d' % ListElement_ItemNumber].
            else:
                raise ValueError('MDWorksheetWorker_DataEditor_ValueRaiseLowCaller Error: Data Passed is unable to met conditions with the code.')
        except ValueError:
            #self.root.ids.
            if TypeAction_ValueChange == 'Increment':
                float(Placement_Val_Conversion)
            elif TypeAction_ValueChange == 'Decrement':
                float(Placement_Val_Conversion)
            else:
                raise ValueError('MDWorksheetWorker_DataEditor_ValueRaiseLowCaller Error: Data Passed is unable to met conditions with the code.')
    def MDWorksheetWorker_DataEditor_CleanInput(self):

        self.root.ids.Resource_ItemNumVal.text = ''
        self.root.ids.Resource_ParticularProperty.text = ''
        self.root.ids.Resource_OnHandVal.text = ''
        self.root.ids.Resource_ProposedVal.text = ''
        self.root.ids.Resource_UnitTypeProperty.text = ''
        self.root.ids.Resource_UnitVal.text = ''
        self.root.ids.Total_ComputedVal.text = ''
        self.root.ids.Quartile_TextFieldVal_One.text = ''
        self.root.ids.Quartile_TextFieldVal_Two.text = ''
        self.root.ids.Quartile_TextFieldVal_Three.text = ''
        self.root.ids.Quartile_TextFieldVal_Four.text = ''
        self.MDUserNotif_SnackbarHandler('Text Field Data has been cleared!')
    
    #Reference this function on the funciton who handles when selected this one it should show up on the data editor
    def IMD_EXPO_ButtonCallBackManager(self, params_WidgetID):
        if params_WidgetID == 'ClearInput_Path':
            self.root.ids.DataBind_FilePath.text = ''
            self.root.ids.DataBind_FileArguments.text = ''
            self.MDUserNotif_SnackbarHandler('Excel File Path and Arguments Input has been cleared!')
        elif params_WidgetID == 'CallFunc_ClearColumnStartPoint':
            self.root.ids.DataBind_ColumnCellStartPoint.text = ''
        elif params_WidgetID == 'CallFunc_ClearColumnEndPoint':
            self.root.ids.DataBind_ColumnCellEndPoint.text = ''
        elif params_WidgetID == 'CallFunc_ClearRowStartPoint':
            self.root.ids.DataBind_RowCellStartPoint.text = ''
        elif params_WidgetID == 'CallFunc_ClearRowEndPoint':
            self.root.ids.DataBind_RowCellEndPoint.text = ''
        elif params_WidgetID == 'CallFunc_ExportProcess':
            pass
        elif params_WidgetID == 'CallFunc_ExportPathClear':
            self.root.ids.DataBind_FileExportPath.text = ''
            self.MDUserNotif_SnackbarHandler('Export Path Input has been cleared!')
        elif params_WidgetID == 'ClearInput_ExcelFileSheet':
            self.root.ids.DataBind_ExcelFileSheet.text = ''
            self.MDUserNotif_SnackbarHandler('Excel Sheet Name Input has been cleared!')
        else:
            raise ValueError('Parameter Variable -> params_TextFieldID: Received a string that is validated with no conditions met.')
            LoggerDebug.error('Parameter Variable -> params_TextFieldID: Received a string that is validated with no conditions met.')
            app.stop()
    #def MDDropdown_CheckAvailSheets(self):
    #    ExcelFile_Lookup = load_workbook(self.root.ids.DataBind_FilePath.text)
    #    for SheetNames in ExcelFile_Lookup.sheetnames:

    def handledrops(self, *args):
        # this will execute each function from list with arguments from
        # Window.on_dropfile
        #
        # make sure `Window.on_dropfile` works on your system first,
        # otherwise the example won't work at all
        for func in self.drops:
            func(*args)

class AvatarSampleWidget(ILeftBody, Image):
    pass

class IconLeftSampleWidget(ILeftBodyTouch, MDIconButton):
    pass

class IconRightSampleWidget(IRightBodyTouch, MDCheckbox):
    pass

if __name__ == '__main__':
    if MD_InventoryEXI_SMV_GUI().CriticalComponent_Check():
        LoggerDebug.info('Resource: File Checks Completed... ')
        MD_InventoryEXI_SMV_GUI().KivyConfig_Setup_FirstTime()
        MD_InventoryEXI_SMV_GUI().run()
    else:
        LoggerDebug.error('CHECK FAILURE: Application was succesfully failed to comply on one of the files to create or check those files. Set the application run with admin previledges...')
        raise Exception('CHECK FAILURE: Application was succesfully failed to comply on one of the files to create or check those files. Set the application run with admin previledges...')
        get_running_app().stop()